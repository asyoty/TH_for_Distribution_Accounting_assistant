import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import xlwings as xw
from datetime import datetime
import os
import sys
import time

# Get the directory where the script or executable is located
if getattr(sys, 'frozen', False):  # If running as a compiled .exe
    script_directory = os.path.dirname(sys.executable)
else:
    script_directory = os.path.dirname(os.path.abspath(__file__))

# List all files in the directory
all_files = os.listdir(script_directory)

# Filter for Excel files with ".xlsx" extension
excel_files = [file for file in all_files if file.endswith('.xlsx')]

while True:
    if len(excel_files) == 1:
        excel_sheet_name = excel_files[0]
        print("Excel sheet detected.")
        break
    elif len(excel_files) > 1:
        print("Multiple Excel sheets found, you should remove the extra sheet(s).")
    else:
        print("No Excel sheet found, you must place the Excel sheet inside the folder.")

    print('App closing in 5 seconds.')
    time.sleep(5)
    sys.exit()



# Load the workbook
workbook = openpyxl.load_workbook(excel_sheet_name)
# Get the target sheet
sheet = workbook['Sales Log Ar']
#save the file
workbook.save(f"{excel_sheet_name}")

## Unmerge all cells in the sheet

# Create a list of merged cell ranges
merged_ranges_list = list(sheet.merged_cells.ranges)

# Iterate over the list and unmerge each cell range
for merged_range in merged_ranges_list:
    sheet.unmerge_cells(str(merged_range))

# Delete the specified number of rows from the sheet
sheet.delete_rows(1, 11)
### delete unwanted columns ###

def find_columns_with_words(search_words_first_row, search_words_third_row):
    # Get the maximum column index
    max_column_index = sheet.max_column
    # Initialize a list to store the column indices for all search words
    List_Of_Columns_To_Keep = []
    # Loop through the first row and find the columns that contain the search words from the first row
    for col in range(1, max_column_index + 1):
        if sheet.cell(row=1, column=col).value in search_words_first_row:
            List_Of_Columns_To_Keep.append(col)
    # Loop through the third row and find the columns that contain the search words from the third row
    for col in range(1, max_column_index + 1):
        if sheet.cell(row=3, column=col).value in search_words_third_row:
            List_Of_Columns_To_Keep.append(col)
    return List_Of_Columns_To_Keep
# Call the function to find columns with specific words in the first row and third row
search_words_first_row = ['جهة التعامل']
search_words_third_row = ['الكمية', 'القيمة', 'الصنف']
List_Of_Columns_To_Keep = find_columns_with_words(search_words_first_row, search_words_third_row)
# Delete all columns from the 'Sales Log Ar' sheet, except for columns with indices in List_Of_Columns_To_Keep
columns_to_delete = [col for col in range(1, sheet.max_column + 1) if col not in List_Of_Columns_To_Keep]
for col in reversed(columns_to_delete):
    sheet.delete_cols(col)

### Loop through and fill the second column of the sheet ###

# Initialize a variable to store the previously copied value
copied_value = None
def find_nearest_non_empty_cell_above(sheet, row, column):
    # Search for the nearest non-empty cell in the specified column, starting from the current row and going upwards
    for row_idx in range(row - 1, 0, -1):
        cell = sheet.cell(row=row_idx, column=column)
        if cell.value is not None:
            return cell.value
    return None
# Loop through the second column of the sheet
for row_idx, row in enumerate(sheet.iter_rows(min_col=2, max_col=2), start=1):
    for cell in row:
        if cell.value is not None:
            # If the cell has a value, update the copied_value
            copied_value = cell.value
        else:
            # If the cell is empty, paste the value from the nearest non-empty cell above
            nearest_value = find_nearest_non_empty_cell_above(sheet, row_idx, cell.column)
            cell.value = nearest_value
# copy a cell to another cell
sheet['B3'] = sheet['B1'].value
# Delete the specified number of rows from the sheet
sheet.delete_rows(1, 2)

### fix the second column ###

# add a new column
sheet.insert_cols(3)
workbook.save(f"{excel_sheet_name}")

# Loop through each row and split the values in the second column using the delimiter "/" (text to columns)

# Create a list to store the modified rows
modified_rows = []

# Loop through each row and split the values in the second column using the delimiter "/"
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[1] is not None:  # Check if the second column is not empty
        split_values = row[1].split('/')
        # If the split_values list has less than 3 elements, fill the remaining elements with None
        row_values = split_values + [None] * (3 - len(split_values))
        # Append the modified row to the list
        modified_rows.append((row[0],) + tuple(row_values))

# Write the modified rows back to the sheet
for i, row in enumerate(modified_rows, start=2):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

#delete the extra column
sheet.delete_cols(3)
#delete last row
sheet.delete_rows(sheet.max_row)

### delete empty rows ###

rows_to_delete = []

# Iterate through rows
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    cell_value = row[0].value
    if cell_value is None:
        rows_to_delete.append(row[0].row)

# Delete rows in reverse order to avoid changing row indexes
for row_num in reversed(rows_to_delete):
    sheet.delete_rows(row_num)

### delete extra rows ###

new_rows_to_delete = []

# Iterate through rows, skipping the first row
for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
    if row[0].value == 'القيمة':
        new_rows_to_delete.append(row)

# Delete the rows in reverse order to avoid shifting indexes
for row in reversed(new_rows_to_delete):
    sheet.delete_rows(row[0].row)

### pivot table ###

# swap first and 4th coulmns
worksheet = workbook.active
# swap first and fourth coulmn
for row in worksheet.iter_rows():
    row[1 - 1].value, row[4 - 1].value = row[4 - 1].value, row[1 - 1].value

workbook.save(f"{excel_sheet_name}")


# add each employee to the top
unique_values = []
for row in workbook.active.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    if row[0] not in unique_values:
        unique_values.append(row[0])
# insert new row
sheet.insert_rows(2)

# Insert new columns and copy values
column_pairs_to_merge = []  # Initialize a list to store pairs of columns for merging

for value in unique_values:
    col_index = workbook.active.max_column - 1  # Insert columns before the third column
    workbook.active.insert_cols(col_index, amount=2)

    # Set header value (without merging cells)
    workbook.active.cell(row=1, column=col_index, value=value)
    workbook.active.cell(row=1, column=col_index + 1, value=value)

    # Store the pairs of cells in the first row for merging later
    column_pairs_to_merge.append((workbook.active.cell(row=1, column=col_index), workbook.active.cell(row=1, column=col_index + 1)))

    # Store the information needed for merging
    col_indices_to_merge = (col_index, col_index + 1)
    col_headers_to_merge = (value, value)

    # Set Arabic header values in the second row
    workbook.active.cell(row=2, column=col_index, value='الكمية')
    workbook.active.cell(row=2, column=col_index + 1, value='القيمة')
# Save the modified Excel file
workbook.save(f"{excel_sheet_name}")

# Load the Excel file
workbook = openpyxl.load_workbook(f"{excel_sheet_name}")
ws = workbook.active

# List to store filtered cells based on conditions
cells_to_fill_1 = []

# Loop through each row except the first two rows
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    for cell in row:
        # Exclude the first two columns and the last two columns
        if cell.column >= 3 and cell.column <= ws.max_column - 2:
            column_cells = [row[cell.column - 1] for row in ws.iter_rows(min_row=3, max_row=ws.max_row)]
            has_value_1 = any(cell.value == 'الكمية' for cell in column_cells)
            has_value_2 = any(cell.value == 'القيمة' for cell in column_cells)

            # Check if the cell in the second row of the column is "الكمية" or "القيمة"
            second_row_cell = ws.cell(row=2, column=cell.column)
            if second_row_cell.value == 'الكمية' and cell.value != 'الكمية':
                has_value_1 = True
            if second_row_cell.value == 'القيمة' and cell.value != 'القيمة':
                has_value_2 = True

            if has_value_1 and cell.value != 'الكمية':
                cells_to_fill_1.append(cell)
            if has_value_2 and cell.value != 'القيمة':
                cells_to_fill_1.append(cell)

# List to store filtered cells based on new exception
cells_to_fill_2 = []

for cell_to_fill in cells_to_fill_1:
    column_letter = cell_to_fill.column_letter
    value_y = ws[f'{column_letter}1'].value
    value_x = ws.cell(row=cell_to_fill.row, column=2).value
    if value_x == value_y:
        cells_to_fill_2.append(cell_to_fill)

# Loop through each cell in cells_to_fill_2
for cell_to_fill in cells_to_fill_2:
    column_cells = [row[cell_to_fill.column - 1] for row in ws.iter_rows(min_row=3, max_row=ws.max_row)]
    has_value_1 = any(cell.value == 'الكمية' for cell in column_cells)
    has_value_2 = any(cell.value == 'القيمة' for cell in column_cells)

    # Check if the cell in the second row of the column is "الكمية" or "القيمة"
    second_row_cell = ws.cell(row=2, column=cell_to_fill.column)
    if second_row_cell.value == 'الكمية' and cell_to_fill.value != 'الكمية':
        has_value_1 = True
    if second_row_cell.value == 'القيمة' and cell_to_fill.value != 'القيمة':
        has_value_2 = True

    if has_value_1:
        value_to_copy = ws.cell(row=cell_to_fill.row, column=ws.max_column - 1).value
        cell_to_fill.value = value_to_copy

    if has_value_2:
        value_to_copy = ws.cell(row=cell_to_fill.row, column=ws.max_column).value
        cell_to_fill.value = value_to_copy
workbook.save(f"{excel_sheet_name}")

workbook = openpyxl.load_workbook(f"{excel_sheet_name}")
# Get the target sheet
sheet = workbook['Sales Log Ar']

sheet.delete_cols(2)
workbook.save(f"{excel_sheet_name}")

### add together the first column

# Create a dictionary to store row sums based on the first column values
row_sums = {}

# Iterate through the rows in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    first_column_value = row[0]
    rest_of_row = row[1:]  # Exclude the first column value

    if first_column_value in row_sums:
        # If the value already exists in the dictionary, update the row sums
        for i, cell_value in enumerate(rest_of_row):
            if cell_value is not None:  # Check if the cell value is not None
                if row_sums[first_column_value][i] is not None:  # Check if the existing value is not None
                    row_sums[first_column_value][i] += cell_value
                else:
                    row_sums[first_column_value][i] = cell_value
    else:
        # If the value doesn't exist, add a new entry to the dictionary
        row_sums[first_column_value] = list(rest_of_row)

# Clear the sheet except for the header row
sheet.delete_rows(2, sheet.max_row - 1)

# Populate the sheet with the updated values from the dictionary
for value, row_sum in row_sums.items():
    sheet.append([value] + row_sum)  # Add back the first column value

### Calculate the total 'الاجمالي'
# Get the last row number
last_row = sheet.max_row

# Add a new row at the end
sheet.append([])

# Set the value in the first column of the new row
cell = sheet.cell(row=last_row + 1, column=1)  # Adjusted the row number to insert a blank row
cell.value = 'الاجمالي'

# Sum up the values for each column and set them in the new row
for col in range(2, sheet.max_column + 1):  # Start from the second column
    col_letter = openpyxl.utils.get_column_letter(col)
    sum_formula = f"SUM({col_letter}2:{col_letter}{last_row})"
    cell = sheet.cell(row=last_row + 1, column=col)
    cell.value = f"={sum_formula}"

#merge header cells
last_col = sheet.max_column
# Merge cells in the first and second rows for the first two and last two columns
workbook.active.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
workbook.active.merge_cells(start_row=1, start_column=last_col - 1, end_row=2, end_column=last_col - 1)
workbook.active.merge_cells(start_row=1, start_column=last_col, end_row=2, end_column=last_col)

##### formatting #####

### arrange the rows alphabetically

# Read all rows into a list (excluding first 2 and last row)
rows = [row for idx, row in enumerate(sheet.iter_rows(), start=1) if idx > 2 and idx < sheet.max_row]

# Sort the rows based on values in the first column (model name) and then the second column (capacity)
sorted_rows = sorted(rows, key=lambda row: (row[0].value.lower(), row[1].value))

# Create a new list with first 2 rows and sorted rows
new_rows = [sheet[1], sheet[2]] + sorted_rows + [sheet[sheet.max_row]]

# Clear the existing content in the sheet
sheet.delete_rows(3, sheet.max_row - 2)

# Write the new rows back to the sheet
for idx, row in enumerate(new_rows, start=1):
    for col_idx, cell in enumerate(row, start=1):
        new_cell = sheet.cell(row=idx, column=col_idx, value=cell.value)

#unfreeze panes
sheet.freeze_panes = None
#row width to 20px
for row in sheet.iter_rows():
    sheet.row_dimensions[row[0].row].height = 45
##columns width

#first column
sheet.column_dimensions['A'].width = 42

#rest of the columns
for i, column in enumerate(sheet.columns):
    if i % 2 != 0:  # Adjust every other column (odd index)
        column_letter = column[0].column_letter
        sheet.column_dimensions[column_letter].width = 8

for i, column in enumerate(sheet.columns):
    if column[0].column_letter != 'A':  # Exclude the first column
        if i % 2 == 0:  # Adjust every other column (even index)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = 17
            for cell in column:
                cell.number_format = '#,##0.00'

# Create font object for Calibri, bold, size 14
bold_font = Font(name='Calibri', bold=True, size=18)

# Create alignment object for center alignment and text wrap
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

# Loop through all cells in the sheet and apply formatting
for row in sheet.iter_rows():
    for cell in row:
        # Apply font formatting
        cell.font = bold_font
        # Apply alignment formatting
        cell.alignment = center_alignment

## merge the pair of cells of the coulmns created using unique_values
for value in range (len(unique_values)):
    if value == 0:
        sheet.merge_cells('B1:C1')
    if value == 1:
        sheet.merge_cells('D1:E1')
    if value == 2:
        sheet.merge_cells('F1:G1')
    if value == 3:
        sheet.merge_cells('H1:I1')
    if value == 4:
        sheet.merge_cells('J1:K1')
    if value == 5:
        sheet.merge_cells('L1:M1')

workbook.save(f"{excel_sheet_name}")

### borders

# Apply the specific border styles to the cells
for row_idx, row in enumerate(sheet.iter_rows(), start=1):
    for col_idx, cell in enumerate(row, start=1):
        if row_idx <= 2 or row_idx > len(list(sheet.iter_rows())) - 1:
            cell.border = Border(
                left=Side(style="thick" if col_idx == 1 else "medium"),
                right=Side(style="thick" if col_idx == len(row) else "medium"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )
        else:
            cell.border = Border(
                left=Side(style="thick" if col_idx == 1 else "medium"),
                right=Side(style="thick" if col_idx == len(row) else "medium"),
                top=Side(style="hair", color="000000"),
                bottom=Side(style="hair", color="000000")
            )

### color the cells

# Define the slightly lighter grey fill color (RGB: 160, 160, 160)
light_grey_fill = PatternFill(start_color='FFA0A0A0', end_color='FFA0A0A0'
                              , fill_type='solid')

# Color the first two rows
for row in sheet.iter_rows(min_row=1, max_row=2):
    for cell in row:
        cell.fill = light_grey_fill

# Color the last row
last_row = sheet.max_row
for cell in sheet[last_row]:
    cell.fill = light_grey_fill

### create new sheet

# Specify the source worksheet
source_worksheet = workbook['Sales Log Ar']

# Create a new worksheet by copying the source worksheet
new_worksheet = workbook.copy_worksheet(source_worksheet)

# delete the old sheet
workbook.remove(source_worksheet)

# Rename the new worksheet (optional)
new_worksheet.title = 'Sales Log Ar'

# Save the modified workbook
workbook.save(f"{excel_sheet_name}")

#
sheet = workbook['Sales Log Ar']
###ready to print (page setup)

# Set the  margins
sheet.page_margins = PageMargins(top= 0.75 , header=0.3, bottom= 0.75,
                                 footer=0.3,right = 0.25, left = 0.25)

## header

# Get the current day of the week and convert it to Arabic
days_in_arabic = {
    0: "الاثنين",
    1: "الثلاثاء",
    2: "الأربعاء",
    3: "الخميس",
    4: "الجمعة",
    5: "السبت",
    6: "الأحد"
}

current_day_index = datetime.now().weekday()
Day_of_the_week = days_in_arabic[current_day_index]
current_date = datetime.now().strftime("%d/%m/%Y")  # Format as "day/month/year"
header_text = f"مبيعات يوم {Day_of_the_week} الموافق {current_date}"

# Assign the header text to the sheet
sheet.oddHeader.center.text = header_text
sheet.evenHeader.center.text = header_text

sheet.oddHeader.center.size = 20
sheet.oddHeader.center.font = "Tahoma,Bold"

sheet.evenHeader.center.size = 20
sheet.evenHeader.center.font = "Tahoma,Bold"

# Assign the footer text to the sheet#

sheet.oddFooter.center.text = "مدير الحسابات"
sheet.oddFooter.center.size = 14
sheet.oddFooter.center.font = "Tahoma,Bold"

sheet.oddFooter.right.text = "    اعداد"
sheet.oddFooter.right.size = 14
sheet.oddFooter.right.font = "Tahoma,Bold"

#this is here to prevent a weird error that ignores formatting for footer center
sheet.oddFooter.left.text = "    "
sheet.oddFooter.left.size = 14
sheet.oddFooter.left.font = "Tahoma,Bold"

# Clear the print area
sheet.print_area = ''

# Save the workbook after making these changes
workbook.save(f"{excel_sheet_name}")


#close openpyxl to use xlwings
workbook.close()

## Adjust the page setup for printing xlwings

# Open the Excel file
wbxw = xw.Book(f"{excel_sheet_name}")

# Choose the sheet you want to modify
sheet = wbxw.sheets['Sales Log Ar']

# Set horizontal centering
sheet.api.PageSetup.CenterHorizontally = True

# Set "Fit to Page" option
sheet.api.PageSetup.Zoom = False
sheet.api.PageSetup.FitToPagesWide = 1
sheet.api.PageSetup.FitToPagesTall = 1

# Save the modified Excel file
wbxw.save()
wbxw.close()


# Load the workbook again
workbook = openpyxl.load_workbook(excel_sheet_name)
# Get the target sheet
sheet = workbook['Sales Log Ar']






# Save the workbook after making these changes
workbook.save(f"{excel_sheet_name}")












# Save the workbook after making these changes
openpyxl.worksheet.page.PrintOptions(horizontalCentered=True)
sheet.page_setup.horizontalCentered = True

workbook.save(f"{excel_sheet_name}")
print('>>>Saving...')





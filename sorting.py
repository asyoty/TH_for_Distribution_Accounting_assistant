import openpyxl
import sys
import os
import time




# Get the directory where the script or executable is located
if getattr(sys, 'frozen', False):  # If running as a compiled .exe
    script_directory = os.path.dirname(sys.executable)
else:
    script_directory = os.path.dirname(os.path.abspath(__file__))

# List all files in the directory
all_files = os.listdir(script_directory)

# Filter for Excel files with ".xlsx" extension, excluding "Order_Database.xlsx"
excel_files = [file for file in all_files if file.endswith('.xlsx') and file != 'Order_Database.xlsx']
while True:
    if len(excel_files) == 1:
        excel_sheet_name = excel_files[0]
        print("Excel sheet detected.")
        break
    elif len(excel_files) > 1:
        print("Multiple Excel sheets found, you should remove the extra sheet(s).")
    else:
        print("No Excel sheet found, you must place the Excel sheet inside the folder.")

    print('App closing in 5 seconds.')
    time.sleep(5)
    sys.exit()
print(excel_files)

### arrange the rows


# Load the workbook
workbook = openpyxl.load_workbook(excel_sheet_name)
sheet = workbook['Sales Log Ar']

# Load the 'Order_Database'
Order_Workbook = openpyxl.load_workbook('Order_Database.xlsx')
order_sheet = Order_Workbook['Item List Report Ar']

# Extract data from both sheets
data = []

# Iterate through the first column (excluding the first two cells)
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=3):
    cell_value = row[0]
    data.append(cell_value)

print("data before:")
print(data)

order_data = []

# Iterate through the first column of the order_sheet
for row in order_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    cell_value = row[0]
    order_data.append(cell_value)


# Assuming 'data' and 'order_data' lists have already been created

# Create a dictionary to store the position of each item in 'order_data'
order_positions = {item: index for index, item in enumerate(order_data)}

# Sort the 'data' list using a custom key function based on 'order_positions'
data.sort(key=lambda item: order_positions.get(item, float('inf')))

# Now 'data' is sorted in the same order as 'order_data'
print("data after:")
print(data)

# Save the modified workbook
workbook.save(excel_sheet_name)

# Save the workbook
workbook.save(excel_sheet_name)

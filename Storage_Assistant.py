import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import xlwings as xw
from datetime import datetime
import os
import sys
import time
import locale


# Get the directory where the script or executable is located
if getattr(sys, 'frozen', False):  # If running as a compiled .exe
    script_directory = os.path.dirname(sys.executable)
else:
    script_directory = os.path.dirname(os.path.abspath(__file__))

# List all files in the directory
all_files = os.listdir(script_directory)

# Filter for Excel files with ".xlsx" extension, excluding "Order_Database.xlsx"
excel_files = [file for file in all_files if file.endswith('.xlsx') and file != 'Order_Database.xlsx']
while True:
    if len(excel_files) == 1:
        excel_sheet_name = excel_files[0]
        print("Excel sheet detected.")
        break
    elif len(excel_files) > 1:
        print("ERROR: Excel sheet is open OR multiple excel sheets are in the folder, you should close the sheet OR you should remove the extra sheet(s).")
    elif len(excel_files) < 1:
        print("ERROR: No Excel sheet found, you must place the Excel sheet inside the folder.")
    else:
        print('ERROR')

    print('App closing in 5 seconds.')
    time.sleep(5)
    sys.exit()

# Load the workbook
workbook = openpyxl.load_workbook(excel_sheet_name)
# Get the target sheet
sheet = workbook['Item Balances Ar']
#save the file
workbook.save(excel_sheet_name)

## Unmerge all cells in the sheet

# Create a list of merged cell ranges
merged_ranges_list = list(sheet.merged_cells.ranges)

# Iterate over the list and unmerge each cell range
for merged_range in merged_ranges_list:
    sheet.unmerge_cells(str(merged_range))

# Delete the specified number of rows from the sheet
sheet.delete_rows(1, 13)

# Delete extra coulumns

[sheet.delete_cols(col_idx) for col_idx in [11,10,9,8, 6, 5, 4, 2]]

#unfreeze panes
sheet.freeze_panes = None

### pivot table ###

#add each storage to the first row

unique_values = []
for row in workbook.active.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
    cell_value = row[0]
    if cell_value is not None and cell_value not in unique_values:
        unique_values.append(cell_value)

# Define the desired order
'''desired_order = ['حازم محمد','حسن فاروق احمد','سامح ناجى','محمد عبدالخالق','سكاى',
'سكاي 2','صيانة ','الرئيسى','المخزن الرئيسي 2022']'''

desired_order = ['حازم محمد', 'حسن فاروق', 'سامح ناجى', 'محمد عبدالخالق', 'سكاى', 'سكاي 2', 'صيانة', 'الرئيسى', 'المخزن الرئيسي 2022']

# Create a dictionary to map each item to its desired position
order_mapping = {item: index for index, item in enumerate(desired_order)}

# Sort the unique_values based on the order_mapping dictionary
sorted_list = sorted(unique_values, key=lambda x: order_mapping.get(x, len(desired_order)))

# Append values to the first row
start_column = 4
for value in unique_values:
    sheet.cell(row=1, column=start_column, value=value)
    start_column += 1  # Move to the next column

workbook.save(excel_sheet_name)

